from wsparser import WordstatParser
import time
import openpyxl


def synonymizer(phrases):
    url = 'https://api-sandbox.direct.yandex.ru/v4/json/'
    # Указываем свой токен на доступ к API Яндекс.Директ
    token = 'y0_AgAAAABorG-EAAklvgAAAADcXOUN2VpAsdrsSeO7Mi-484B6xHYGioI'
    userName = 'pyti55'

    # Пишем список общих минус-слов, как в примере (со знаком "-")
    minusWords = [
    ]
    # Код скрипта парсинга
    geo = []
  
    # Добавляем минус-слова ко всем фразам
    data = []  
    

    
    for i in range(len(phrases)):
        data.append(phrases[i])
        for j in range(len(minusWords)):
            data[i] += ' '+minusWords[j]

    # Создаем парсер
    parser = WordstatParser(url, token, userName)

    try:
        # Запрашиваем кол-во оставшихся баллов Яндекс.Директ API
        units = parser.getClientUnits()
        if 'data' in units:
            print ('>>> Баллов осталось: ', units['data'][0]['UnitsRest'])
        else:
            raise Exception('Не удалось получить баллы', units)

        # Отправляем запрос на создание нового отчета на сервере Яндекс.Директ
        response = parser.createReport(data, geo)
        if 'data' in response:
            reportID = response['data']
            print ('>>> Создается отчет с ID = ', reportID)
        else:
            raise Exception('Не удалось создать отчет', response)
            
        # Проверяем список отчетов на сервере. Должен появиться новый. Ожидаем его готовности
        reportList = parser.getReportList()
        if 'data' in reportList:
            lastReport = reportList['data'][len(reportList['data'])-1]
            i = 0
            while lastReport['StatusReport'] != 'Done':
                print ('>>> Подготовка отчета, ждите ... ('+str(i)+')')
                time.sleep(2)
                reportList = parser.getReportList()
                lastReport = reportList['data'][len(reportList['data'])-1]
                i+=1
            print ('>>> Отчет ID = ', lastReport['ReportID'], ' получен!')
        else:
            raise Exception('Не удалось прочитать список отчетов', reportList)

        # Читаем отчет
        report = parser.readReport(reportID)
        if 'data' in report:
            # Сохраняем результаты парсинга в файлы (отдельно фразы, отдельно частотности). 
            # Если rightCol == True, будет сохраняться правая колонка Яндекс.Вордстат (в дополнение к левой)
            parser.saveReportToTxt(report, True)
            print ('>>> Результаты парсига успешно сохранены в файлы!')
        else:
            raise Exception('Не удалось прочитать отчет', report)
        
        # Удаляем на сервере Яндекс.Директ новый отчет (он больше не нужен)
        report = parser.deleteReport(reportID)
        if 'data' in report:
            print ('>>> Отчет с ID = ', reportID, ' успешно удален с сервера Яндекс.Директ')
        else:
            raise Exception('Не удалось удалить отчет', report)

        print ('>>> Все готово!')

    except Exception as e:
        print ('>>> Поймано исключение:', e)
 

def read_excel():
    a_list = []
    num = 1
    num_a = 1

    with open('phrases_left.txt', 'r') as files:
        for file in files:
            a_list.append(file)


    wb = openpyxl.Workbook() 
    sheet = wb.active 

    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "Hello"

    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "World"

    wb.save("ТОП2.xlsx") 

    for l in a_list:
        a = sheet['A' + str(num)] 
        a.value = num_a 

        c = sheet['B' + str(num)] 
        c.value = l + ' IEK'
        wb.save("ТОП2.xlsx")

        num_a += 1
        num += 1



if __name__ == '__main__':
    excel_name = input('Введите название excel файла:')
    wb = openpyxl.load_workbook(excel_name)
    phrases = []
    sheets = wb.sheetnames
    sheet = wb.active
    sheet_count = wb.worksheets[0]
    row_count = sheet_count.max_row


    for n in range(1, row_count):
        words = sheet['B' + str(n)].value
        a = words.replace('IEK', '').replace('.', '').replace(',', '').replace('=', '').replace('-', '').replace('/', '').replace('"', '').replace('(', '').replace(')', '').replace('+', '').replace(':', '').replace(';', '').split()[:7]
        phrases.append(' '.join(i for i in a))
    

    n = 0
    n_1 = 0
    n_2 = 10
    num_l = len(phrases) // 10

    for i in range(num_l):
        print(synonymizer(phrases[n_1:n_2]))
        n_1 += 10
        n_2 += 10

print(read_excel())
